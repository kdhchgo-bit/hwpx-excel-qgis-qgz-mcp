from __future__ import annotations

import os
import re
from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries

from .common import copy_for_edit, output_copy_path, require_file

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _load_workbook(path: Path, data_only: bool = False, read_only: bool = False):
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    return openpyxl.load_workbook(path, data_only=data_only, read_only=read_only, keep_vba=keep_vba, keep_links=True)


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _rectangular(values: list[list[Any]]) -> tuple[int, int]:
    if not values or not values[0]:
        raise ValueError("values must be a non-empty two-dimensional array")
    width = len(values[0])
    if any(len(row) != width for row in values):
        raise ValueError("values must be rectangular")
    return len(values), width


def excel_health() -> dict[str, Any]:
    """Return Excel package and native COM availability."""
    com_available = False
    error = None
    try:
        import win32com.client  # noqa: F401

        com_available = True
    except Exception as exc:  # pragma: no cover - platform dependent
        error = str(exc)
    return {
        "server": "excel-local",
        "openpyxl_version": openpyxl.__version__,
        "excel_com_available": com_available,
        "excel_com_error": error,
        "source_overwrite": False,
        "features": ["inspect", "read_range", "find", "write_range", "recalculate", "export_pdf"],
    }


def excel_inspect(path: str, include_defined_names: bool = True) -> dict[str, Any]:
    """Inspect Excel sheet dimensions, visibility, merged cells, tables, and formulas."""
    source = require_file(path, EXCEL_SUFFIXES)
    workbook = _load_workbook(source, data_only=False, read_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            formula_count = 0
            nonempty_count = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        nonempty_count += 1
                        if cell.data_type == "f":
                            formula_count += 1
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "dimension": sheet.calculate_dimension(),
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "nonempty_cells": nonempty_count,
                    "formula_cells": formula_count,
                    "merged_ranges": [str(item) for item in list(sheet.merged_cells.ranges)[:500]],
                    "table_names": list(sheet.tables.keys()),
                    "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                }
            )
        defined_names = []
        if include_defined_names:
            defined_names = [item.name for item in workbook.defined_names.values()]
        return {
            "path": str(source),
            "bytes": source.stat().st_size,
            "sheet_count": len(sheets),
            "active_sheet": workbook.active.title,
            "sheets": sheets,
            "defined_names": defined_names,
            "vba_preserved_on_write": source.suffix.lower() in {".xlsm", ".xltm"},
        }
    finally:
        workbook.close()


def excel_read_range(
    path: str,
    sheet_name: str,
    cell_range: str,
    data_only: bool = False,
    include_number_formats: bool = False,
) -> dict[str, Any]:
    """Read a rectangular Excel range with formulas or cached values."""
    source = require_file(path, EXCEL_SUFFIXES)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
    if cell_count > 20_000:
        raise ValueError(f"Range is too large ({cell_count} cells); limit is 20,000")
    workbook = _load_workbook(source, data_only=data_only, read_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        values: list[list[Any]] = []
        formats: list[list[str]] = []
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            values.append([_json_value(cell.value) for cell in row])
            if include_number_formats:
                formats.append([cell.number_format for cell in row])
        return {
            "path": str(source),
            "sheet": sheet_name,
            "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            "data_only": data_only,
            "values": values,
            "number_formats": formats if include_number_formats else None,
        }
    finally:
        workbook.close()


def excel_find(
    path: str,
    query: str,
    regex: bool = False,
    sheet_name: str | None = None,
    max_results: int = 200,
    max_scanned_cells: int = 1_000_000,
) -> dict[str, Any]:
    """Find values or formulas across an Excel workbook."""
    if not query:
        raise ValueError("query must not be empty")
    source = require_file(path, EXCEL_SUFFIXES)
    pattern = re.compile(query) if regex else None
    max_results = max(1, min(max_results, 5000))
    workbook = _load_workbook(source, data_only=False, read_only=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
        results: list[dict[str, Any]] = []
        match_count = 0
        scanned = 0
        scan_truncated = False
        for sheet in sheets:
            for row in sheet.iter_rows():
                for cell in row:
                    scanned += 1
                    if scanned > max_scanned_cells:
                        scan_truncated = True
                        break
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    hit = bool(pattern.search(text)) if pattern else query in text
                    if hit:
                        match_count += 1
                        if len(results) < max_results:
                            results.append({"sheet": sheet.title, "cell": cell.coordinate, "value": _json_value(cell.value)})
                if scan_truncated:
                    break
            if scan_truncated:
                break
        return {
            "path": str(source),
            "query": query,
            "match_count": match_count,
            "results": results,
            "results_truncated": match_count > len(results),
            "scanned_cells": scanned,
            "scan_truncated": scan_truncated,
        }
    finally:
        workbook.close()


def _write_openpyxl(
    source: Path,
    output: Path,
    sheet_name: str,
    start_cell: str,
    values: list[list[Any]],
    copy_style_from: str | None,
    overwrite: bool,
) -> None:
    copy_for_edit(source, output, overwrite)
    workbook = _load_workbook(output, data_only=False, read_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        start = sheet[start_cell]
        style_source = sheet[copy_style_from] if copy_style_from and ":" not in copy_style_from else None
        for row_offset, row in enumerate(values):
            for column_offset, value in enumerate(row):
                cell = sheet.cell(row=start.row + row_offset, column=start.column + column_offset)
                if style_source is not None:
                    cell._style = copy(style_source._style)
                    cell.number_format = style_source.number_format
                    cell.protection = copy(style_source.protection)
                    cell.alignment = copy(style_source.alignment)
                cell.value = value
        workbook.save(output)
    finally:
        workbook.close()


def _excel_com_session():
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    app = win32com.client.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    app.AskToUpdateLinks = False
    return pythoncom, app


def _write_com(
    source: Path,
    output: Path,
    sheet_name: str,
    start_cell: str,
    values: list[list[Any]],
    copy_style_from: str | None,
    overwrite: bool,
    recalculate: bool,
) -> None:
    copy_for_edit(source, output, overwrite)
    pythoncom, app = _excel_com_session()
    workbook = None
    try:
        workbook = app.Workbooks.Open(str(output), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
        sheet = workbook.Worksheets(sheet_name)
        start = sheet.Range(start_cell)
        rows, columns = _rectangular(values)
        target = sheet.Range(
            start,
            sheet.Cells(start.Row + rows - 1, start.Column + columns - 1),
        )
        if copy_style_from:
            sheet.Range(copy_style_from).Copy(Destination=target)
        for row_offset, row in enumerate(values):
            for column_offset, value in enumerate(row):
                cell = sheet.Cells(start.Row + row_offset, start.Column + column_offset)
                if isinstance(value, str) and value.startswith("="):
                    cell.Formula = value
                else:
                    cell.Value2 = value
        if recalculate:
            app.CalculateFullRebuild()
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        app.Quit()
        pythoncom.CoUninitialize()


def excel_write_range(
    path: str,
    sheet_name: str,
    start_cell: str,
    values: list[list[Any]],
    output_path: str | None = None,
    copy_style_from: str | None = None,
    engine: str = "com",
    recalculate: bool = True,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Write a rectangular range to a new workbook copy; COM preserves native Excel fidelity."""
    rows, columns = _rectangular(values)
    source = require_file(path, EXCEL_SUFFIXES)
    output = output_copy_path(source, output_path, "MCP수정", overwrite)
    engine = engine.lower()
    if engine == "com":
        if os.name != "nt":
            raise RuntimeError("Excel COM is only available on Windows")
        _write_com(source, output, sheet_name, start_cell, values, copy_style_from, overwrite, recalculate)
    elif engine == "openpyxl":
        _write_openpyxl(source, output, sheet_name, start_cell, values, copy_style_from, overwrite)
    else:
        raise ValueError("engine must be 'com' or 'openpyxl'")
    return {
        "source": str(source),
        "output": str(output),
        "sheet": sheet_name,
        "start_cell": start_cell,
        "rows": rows,
        "columns": columns,
        "cells_written": rows * columns,
        "engine": engine,
        "recalculated": bool(engine == "com" and recalculate),
        "output_bytes": output.stat().st_size,
    }


def excel_recalculate(
    path: str,
    output_path: str | None = None,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Use installed Excel to fully recalculate a workbook and save a separate copy."""
    if os.name != "nt":
        raise RuntimeError("Excel COM is only available on Windows")
    source = require_file(path, EXCEL_SUFFIXES)
    output = output_copy_path(source, output_path, "재계산", overwrite)
    copy_for_edit(source, output, overwrite)
    pythoncom, app = _excel_com_session()
    workbook = None
    try:
        workbook = app.Workbooks.Open(str(output), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
        app.CalculateFullRebuild()
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        app.Quit()
        pythoncom.CoUninitialize()
    return {"source": str(source), "output": str(output), "recalculated": True, "output_bytes": output.stat().st_size}


def excel_export_pdf(
    path: str,
    output_pdf: str,
    sheet_name: str | None = None,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Export an entire workbook or one worksheet to PDF using installed Excel."""
    if os.name != "nt":
        raise RuntimeError("Excel COM is only available on Windows")
    source = require_file(path, EXCEL_SUFFIXES)
    output = Path(output_pdf).expanduser().resolve()
    if output.suffix.lower() != ".pdf":
        raise ValueError("output_pdf must end with .pdf")
    if output.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {output}")
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        output.unlink()
    pythoncom, app = _excel_com_session()
    workbook = None
    try:
        workbook = app.Workbooks.Open(str(source), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
        target = workbook.Worksheets(sheet_name) if sheet_name else workbook
        target.ExportAsFixedFormat(0, str(output))
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        app.Quit()
        pythoncom.CoUninitialize()
    return {"source": str(source), "output_pdf": str(output), "sheet": sheet_name, "bytes": output.stat().st_size}


def register_excel_tools(mcp: Any) -> None:
    for function in (excel_health, excel_inspect, excel_read_range, excel_find, excel_write_range, excel_recalculate, excel_export_pdf):
        mcp.tool()(function)
